from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADERS = [
    "Parça No", "Parça Adı", "Pattern / Kalıp No", "Makine / Pres",
    "Aylık Üretim Adedi", "Parça Başı Süre sn", "Proje", "Hat / Bölge",
    "Müşteri", "Öncelik", "Açıklama",
]

SAMPLE_ROWS = [
    ["PRT-001", "Ön Braket RH", "PTN-1001", "Pres-1", 12000, 8.5, "Proje-A", "W1-1", "Toyota", "Yüksek", "Örnek veri"],
    ["PRT-002", "Ön Braket LH", "PTN-1002", "Pres-1", 11500, 8.7, "Proje-A", "W1-1", "Toyota", "Yüksek", "Örnek veri"],
    ["PRT-003", "Reinforcement Plate", "PTN-2001", "Pres-2", 8500, 14.2, "Proje-B", "W1-2", "Toyota", "Orta", "Örnek veri"],
    ["PRT-004", "Side Member", "PTN-3001", "Pres-3", 6200, 22.5, "Proje-C", "W2-1", "Toyota", "Orta", "Örnek veri"],
    ["PRT-005", "Cross Member", "PTN-3002", "Pres-3", 4500, 31.0, "Proje-C", "W2-1", "Toyota", "Düşük", "Örnek veri"],
    ["PRT-006", "Support Panel", "PTN-4001", "Pres-4", 15000, 6.8, "Proje-D", "W3-1", "Toyota", "Yüksek", "Örnek veri"],
    ["PRT-007", "Inner Plate", "PTN-5001", "Pres-2", 9800, 12.0, "Proje-B", "W1-2", "Toyota", "Orta", "Örnek veri"],
    ["PRT-008", "Outer Plate", "PTN-5002", "Pres-4", 7200, 18.4, "Proje-D", "W3-1", "Toyota", "Düşük", "Örnek veri"],
]

DESCRIPTIONS = [
    ("Parça No", "Üretilecek parçanın benzersiz numarası"),
    ("Parça Adı", "Parçanın açıklayıcı adı"),
    ("Pattern / Kalıp No", "Kullanılacak pattern veya kalıp numarası"),
    ("Makine / Pres", "Üretimin yapılacağı pres veya makine adı"),
    ("Aylık Üretim Adedi", "İlgili ayda üretilmesi gereken toplam adet"),
    ("Parça Başı Süre sn", "1 adet üretim için gereken süre, saniye cinsinden"),
    ("Proje", "İsteğe bağlı proje bilgisi"),
    ("Hat / Bölge", "İsteğe bağlı hat veya bölge bilgisi"),
    ("Müşteri", "İsteğe bağlı müşteri bilgisi"),
    ("Öncelik", "Yüksek / Orta / Düşük gibi öncelik bilgisi"),
    ("Açıklama", "Ek notlar"),
]


def create_data_template(path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Veri Şablonu"

    sheet.append(HEADERS)
    for row in SAMPLE_ROWS:
        sheet.append(row)

    navy = PatternFill("solid", fgColor="101F35")
    red = PatternFill("solid", fgColor="C9182B")
    pale = PatternFill("solid", fgColor="EAF0F7")
    white_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="B8C5D3")

    for cell in sheet[1]:
        cell.fill = red
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=thin)
    sheet.row_dimensions[1].height = 28
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for cell in row:
            cell.fill = pale if cell.row % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
            cell.border = Border(bottom=Side(style="hair", color="D7E0EA"))
            cell.alignment = Alignment(vertical="center")
    for row in range(2, sheet.max_row + 1):
        sheet.cell(row, 5).number_format = "#,##0"
        sheet.cell(row, 6).number_format = "0.0"
    widths = [14, 24, 22, 16, 21, 20, 14, 15, 14, 12, 24]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:K{sheet.max_row}"
    sheet.sheet_view.showGridLines = False

    notes = workbook.create_sheet("Açıklamalar")
    notes.append(["Kolon", "Açıklama"])
    for item in DESCRIPTIONS:
        notes.append(item)
    for cell in notes[1]:
        cell.fill = navy
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center")
    notes.column_dimensions["A"].width = 26
    notes.column_dimensions["B"].width = 68
    notes.freeze_panes = "A2"
    notes.auto_filter.ref = f"A1:B{notes.max_row}"
    notes.sheet_view.showGridLines = False
    for row in notes.iter_rows(min_row=2):
        row[0].font = Font(bold=True, color="C9182B")
        row[1].alignment = Alignment(wrap_text=True, vertical="top")
    workbook.save(path)
